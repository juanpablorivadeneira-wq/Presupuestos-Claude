"""
Generates a .xlsx file with four sheets:
  1. Resumen        – sorted quantity table
  2. Sin datos      – sections with no Revit data
  3. Dashboard      – KPI summary + chapter completeness + top-10 tables
  4. Warnings       – structured warning list
"""

import io
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .analyzer import AnalysisResult, SectionSummary
from .warnings import Warning, WarnLevel, build_warning_list

# ── Colour palette ────────────────────────────────────────────────────────────
_BLUE_DARK = "1F4E78"
_BLUE_MED = "2E75B6"
_GREEN_DARK = "375623"
_RED = "C00000"
_YELLOW = "FFC000"
_LIGHT_BLUE = "D6E4F0"
_LIGHT_GREEN = "E2EFDA"
_LIGHT_RED = "FCE4D6"
_WHITE = "FFFFFF"
_GREY_LIGHT = "F2F2F2"
_GREY_BORDER = "BFBFBF"

_THIN = Side(style="thin", color=_GREY_BORDER)
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_style(bg: str, fg: str = _WHITE, bold: bool = True) -> dict[str, Any]:
    return {
        "fill": PatternFill("solid", fgColor=bg),
        "font": Font(bold=bold, color=fg, size=10),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _THIN_BORDER,
    }


def _apply_style(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row: int, values: list, style: dict | None = None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if style:
            _apply_style(cell, **style)


# ── Sheet 1: Resumen ──────────────────────────────────────────────────────────

def _build_resumen(wb: openpyxl.Workbook, ar: AnalysisResult):
    ws = wb.active
    ws.title = "Resumen"

    headers = ["Capítulo", "Código", "Descripción", "Cantidad", "Unidad", "Fuente", "# Elementos"]
    hstyle = _header_style(_BLUE_DARK)

    _write_row(ws, 1, headers, hstyle)
    ws.row_dimensions[1].height = 28

    number_fmt = "#,##0.00"
    data_rows: list[SectionSummary] = ar.with_data

    for i, s in enumerate(data_rows, 2):
        row_bg = _LIGHT_BLUE if i % 2 == 0 else _WHITE
        row_fill = PatternFill("solid", fgColor=row_bg)
        base_style = {
            "fill": row_fill,
            "font": Font(size=10),
            "alignment": Alignment(vertical="center"),
            "border": _THIN_BORDER,
        }
        cells = [
            f"{s.chapter_code} – {s.chapter_name}",
            s.code,
            s.description,
            s.quantity,
            s.unit,
            s.fuente,
            s.element_count,
        ]
        for col, val in enumerate(cells, 1):
            cell = ws.cell(row=i, column=col, value=val)
            _apply_style(cell, **base_style)
            if col == 4 and isinstance(val, (int, float)):
                cell.number_format = number_fmt
            if col == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if col in (4, 7):
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Unit totals at bottom
    last_data_row = len(data_rows) + 1
    totals_row = last_data_row + 2
    totals = ar.totals_by_unit()
    ws.cell(row=totals_row, column=3, value="TOTALES POR UNIDAD").font = Font(bold=True, size=10)
    for r, (unit, total) in enumerate(totals.items(), totals_row + 1):
        ws.cell(row=r, column=4, value=total).number_format = number_fmt
        ws.cell(row=r, column=5, value=unit)
        ws.cell(row=r, column=4).font = Font(bold=True)

    # Column widths
    widths = [28, 14, 45, 14, 8, 22, 12]
    for col, w in enumerate(widths, 1):
        _set_col_width(ws, col, w)

    ws.freeze_panes = "A2"


# ── Sheet 2: Sin datos ────────────────────────────────────────────────────────

def _build_sin_datos(wb: openpyxl.Workbook, ar: AnalysisResult):
    ws = wb.create_sheet("Sin datos")
    headers = ["Código", "Descripción", "Capítulo", "Observación"]
    hstyle = _header_style(_RED)
    _write_row(ws, 1, headers, hstyle)
    ws.row_dimensions[1].height = 24

    obs = ("Tabla presente en Revit pero sin elementos modelados. "
           "Requiere revisión: modelado pendiente o rubro no aplica.")

    prev_chapter = None
    row = 2
    for s in ar.without_data:
        if s.chapter_code != prev_chapter:
            # Chapter separator row
            ch_cell = ws.cell(row=row, column=1,
                              value=f"  {s.chapter_code} – {s.chapter_name}")
            ch_cell.font = Font(bold=True, color=_WHITE, size=10)
            ch_cell.fill = PatternFill("solid", fgColor=_BLUE_MED)
            ws.merge_cells(f"A{row}:D{row}")
            row += 1
            prev_chapter = s.chapter_code

        row_bg = _LIGHT_RED if row % 2 == 0 else _WHITE
        row_fill = PatternFill("solid", fgColor=row_bg)
        vals = [s.code, s.description, f"{s.chapter_code} – {s.chapter_name}", obs]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(size=10)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))
        row += 1

    widths = [14, 45, 28, 70]
    for col, w in enumerate(widths, 1):
        _set_col_width(ws, col, w)

    ws.freeze_panes = "A2"


# ── Sheet 3: Dashboard ────────────────────────────────────────────────────────

def _build_dashboard(wb: openpyxl.Workbook, ar: AnalysisResult):
    ws = wb.create_sheet("Dashboard")

    # ── KPI cards block ───────────────────────────────────────────────────────
    kpis = [
        ("Rubros totales", ar.total_count),
        ("Con datos", len(ar.with_data)),
        ("Sin datos", len(ar.without_data)),
        ("% Modelado", f"{100 * len(ar.with_data) / ar.total_count:.1f}%" if ar.total_count else "0%"),
    ]
    kpi_colors = [_BLUE_DARK, _GREEN_DARK, _RED, _BLUE_MED]

    ws.cell(row=1, column=1, value="DASHBOARD DE CANTIDADES DE OBRA").font = Font(
        bold=True, color=_BLUE_DARK, size=14
    )
    ws.row_dimensions[1].height = 30

    row = 3
    for col, ((label, val), bg) in enumerate(zip(kpis, kpi_colors), 1):
        lbl_cell = ws.cell(row=row, column=col * 2 - 1, value=label)
        lbl_cell.font = Font(bold=True, color=_WHITE, size=9)
        lbl_cell.fill = PatternFill("solid", fgColor=bg)
        lbl_cell.alignment = Alignment(horizontal="center")

        val_cell = ws.cell(row=row + 1, column=col * 2 - 1, value=val)
        val_cell.font = Font(bold=True, color=bg, size=16)
        val_cell.alignment = Alignment(horizontal="center")

    # ── Unit totals ────────────────────────────────────────────────────────────
    row = 6
    totals = ar.totals_by_unit()
    ws.cell(row=row, column=1, value="TOTALES AGREGADOS").font = Font(
        bold=True, color=_BLUE_DARK, size=11
    )
    row += 1
    for unit, total in totals.items():
        ws.cell(row=row, column=1, value=unit).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=total)
        cell.number_format = "#,##0.00"
        row += 1

    # ── Chapter completeness table ────────────────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="COMPLETITUD POR CAPÍTULO").font = Font(
        bold=True, color=_BLUE_DARK, size=11
    )
    row += 1

    ch_headers = ["Capítulo", "Nombre", "Con datos", "Sin datos", "Total", "% Modelado"]
    for col, h in enumerate(ch_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        _apply_style(cell, **_header_style(_BLUE_DARK))

    row += 1
    for ch_code, stats in ar.chapter_stats().items():
        pct = stats["pct"]
        bg = _LIGHT_GREEN if pct >= 80 else (_LIGHT_RED if pct < 50 else _GREY_LIGHT)
        row_fill = PatternFill("solid", fgColor=bg)
        vals = [ch_code, stats["name"], stats["with_data"], stats["without_data"],
                stats["total"], f"{pct:.1f}%"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(size=10)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")
        row += 1

    # ── Top 10 by area ────────────────────────────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="TOP 10 RUBROS POR ÁREA (m²)").font = Font(
        bold=True, color=_BLUE_DARK, size=11
    )
    row += 1
    top_headers = ["#", "Código", "Descripción", "Área (m²)"]
    for col, h in enumerate(top_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        _apply_style(cell, **_header_style(_BLUE_DARK))

    row += 1
    for rank, s in enumerate(ar.top_by_unit("m²"), 1):
        row_fill = PatternFill("solid", fgColor=_LIGHT_BLUE if rank % 2 == 0 else _WHITE)
        vals = [rank, s.code, s.description, s.quantity]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(size=10)
            cell.border = _THIN_BORDER
            if col == 4:
                cell.number_format = "#,##0.00"
        row += 1

    # Column widths
    widths = [14, 16, 45, 16, 12, 12]
    for col, w in enumerate(widths, 1):
        _set_col_width(ws, col, w)


# ── Sheet 4: Warnings ─────────────────────────────────────────────────────────

def _build_warnings(wb: openpyxl.Workbook, warnings: list[Warning]):
    ws = wb.create_sheet("Warnings")
    headers = ["Código", "Nivel", "Sección", "Mensaje"]
    _write_row(ws, 1, headers, _header_style(_YELLOW, fg="000000"))
    ws.row_dimensions[1].height = 24

    level_colors = {
        WarnLevel.ERROR: _LIGHT_RED,
        WarnLevel.WARNING: "FFF2CC",
        WarnLevel.INFO: _LIGHT_BLUE,
    }

    for i, w in enumerate(warnings, 2):
        bg = level_colors.get(w.level, _WHITE)
        row_fill = PatternFill("solid", fgColor=bg)
        vals = [w.code, w.level.value, w.section_code or "", w.message]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(size=10)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))

    widths = [8, 10, 16, 80]
    for col, w in enumerate(widths, 1):
        _set_col_width(ws, col, w)

    ws.freeze_panes = "A2"

    if not warnings:
        ws.cell(row=2, column=1, value="Sin advertencias").font = Font(
            bold=True, color=_GREEN_DARK
        )


# ── Public API ────────────────────────────────────────────────────────────────

def export_to_bytes(ar: AnalysisResult) -> bytes:
    """Generate the Excel workbook and return its raw bytes."""
    wb = openpyxl.Workbook()

    _build_resumen(wb, ar)
    _build_sin_datos(wb, ar)
    _build_dashboard(wb, ar)

    structured_warnings = build_warning_list(ar.warnings)
    _build_warnings(wb, structured_warnings)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
