"""Tests for the Excel exporter."""

import io
import pytest
import openpyxl
from pathlib import Path

from ..parsers.revit_csv import parse
from ..services.analyzer import analyze
from ..services.exporter import export_to_bytes


FIXTURE = Path(__file__).parent / "fixtures" / "Todas_las_Tablas.csv"


@pytest.fixture(scope="module")
def xlsx_bytes() -> bytes:
    ar = analyze(parse(FIXTURE))
    return export_to_bytes(ar)


class TestExcelStructure:
    def test_produces_bytes(self, xlsx_bytes):
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 1000

    def test_valid_xlsx(self, xlsx_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb is not None

    def test_four_sheets_present(self, xlsx_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert set(wb.sheetnames) == {"Resumen", "Sin datos", "Dashboard", "Warnings"}


class TestResumenSheet:
    def test_header_row_present(self, xlsx_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Resumen"]
        header = [ws.cell(1, c).value for c in range(1, 8)]
        assert "Código" in header
        assert "Descripción" in header
        assert "Cantidad" in header

    def test_data_rows_present(self, xlsx_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Resumen"]
        # At minimum the sections with data should each have a row
        data_values = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1)
                       if ws.cell(r, 2).value]
        assert len(data_values) >= 10

    def test_quantity_is_numeric(self, xlsx_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Resumen"]
        for row in range(2, ws.max_row + 1):
            qty = ws.cell(row, 4).value
            if qty is not None:
                assert isinstance(qty, (int, float))


class TestSinDatosSheet:
    def test_has_data_rows(self, xlsx_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Sin datos"]
        codes = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)
                 if ws.cell(r, 1).value]
        assert len(codes) >= 5

    def test_observation_text_present(self, xlsx_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Sin datos"]
        obs_texts = [ws.cell(r, 4).value for r in range(2, ws.max_row + 1)
                     if ws.cell(r, 4).value and "modelado" in str(ws.cell(r, 4).value).lower()]
        assert len(obs_texts) >= 1


class TestDashboardSheet:
    def test_kpi_labels_present(self, xlsx_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Dashboard"]
        all_values = [ws.cell(r, c).value for r in range(1, 15) for c in range(1, 10)
                      if ws.cell(r, c).value]
        text_vals = [str(v).upper() for v in all_values]
        assert any("TOTALES" in v for v in text_vals)


class TestWarningsSheet:
    def test_warnings_sheet_has_data(self, xlsx_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Warnings"]
        # Row 2 should have content (either warnings or "Sin advertencias")
        assert ws.cell(2, 1).value is not None
